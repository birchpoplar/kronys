import pandas as pd
import sqlite3
from sqlite3 import Error
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.utils import get_column_letter


def fetch_transactions_to_df(conn):
    """Fetch all transactions from the general ledger table and return as a pandas dataframe."""
    sql_fetch_transactions = "SELECT * FROM general_ledger;"
    return pd.read_sql(sql_fetch_transactions, conn)


def income_statement_dataframe(accounts, transactions, start_month, end_month):
    # Filter Income and Expense accounts
    income_expense_accounts = [account for account in accounts if account.category in ['Income', 'Expense']]

    # Create a dataframe from the transactions with only the needed columns
    transactions_df = transactions[['month', 'account', 'debit', 'credit']]

    # Create an empty dataframe to store the income statement
    income_statement = pd.DataFrame(columns=['account_number', 'account_name', 'type'] + list(range(start_month, end_month+1)))

    for account in income_expense_accounts:
        account_transactions = transactions_df[transactions_df['account'] == account.account_number]
        net_by_month = []

        for month in range(start_month, end_month+1):
            monthly_transactions = account_transactions[account_transactions['month'] == month]
            total_debit = monthly_transactions['debit'].sum()
            total_credit = monthly_transactions['credit'].sum()

            if account.type == 'DR':
                net_amount = total_debit - total_credit
            else:
                net_amount = total_credit - total_debit

            net_by_month.append(net_amount)

        new_row = pd.DataFrame({
            'account_number': [account.account_number],
            'account_name': [account.account_name],
            'type': [account.type],
            **dict(zip(range(start_month, end_month+1), net_by_month))
        })

        income_statement = pd.concat([income_statement, new_row], ignore_index=True)

    return income_statement


def balance_sheet_dataframe(accounts, transactions, start_month, end_month):
    # Filter Assets, Liabilities, and Equity accounts
    balance_sheet_accounts = [account for account in accounts if account.category in ['Assets', 'Liabilities', 'Equity']]

    # Create a dataframe from the transactions with only the needed columns
    transactions_df = transactions[['month', 'account', 'debit', 'credit']]

    # Create an empty dataframe to store the balance sheet
    balance_sheet = pd.DataFrame(columns=['account_number', 'account_name', 'type'] + list(range(start_month, end_month+1)))

    for account in balance_sheet_accounts:
        account_transactions = transactions_df[transactions_df['account'] == account.account_number]
        net_by_month = []
        cumulative_balance = 0

        for month in range(start_month, end_month+1):
            monthly_transactions = account_transactions[account_transactions['month'] == month]
            total_debit = monthly_transactions['debit'].sum()
            total_credit = monthly_transactions['credit'].sum()

            if account.type == 'DR':
                net_amount = total_debit - total_credit
            else:
                net_amount = total_credit - total_debit

            cumulative_balance += net_amount
            net_by_month.append(cumulative_balance)

        new_row = pd.DataFrame({
            'account_number': [account.account_number],
            'account_name': [account.account_name],
            'type': [account.type],
            **dict(zip(range(start_month, end_month+1), net_by_month))
        })

        balance_sheet = pd.concat([balance_sheet, new_row], ignore_index=True)

    return balance_sheet


def create_excel_file(income_statement_df, file_name, start_month, end_month):
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    # Write the header row
    header_row = ['Account Number', 'Account Name', 'Type'] + [f'Month {i}' for i in range(start_month, end_month+1)]
    for col, header in enumerate(header_row, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write the DataFrame rows
    for r_idx, row in enumerate(dataframe_to_rows(income_statement_df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file_name)
    print(f'Income statement saved to {file_name}')


def add_total_row(worksheet, row_num):
    for col_idx in range(2, worksheet.max_column + 1):
        col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{row_num - 1})"
        worksheet[f"{col_letter}{row_num}"] = formula


